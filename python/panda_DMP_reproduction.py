# -*- coding: utf-8 -*-
"""
This code is a 2D simulation of imitation learning for a two-link robotic arm based on DMP.
In this program, the user will draw a demonstration trajectory using the mouse,
and the trained DMP will drive the robotic arm to reproduce or generalize the trajectory.

Author: Cheems_JH
Date: 2023.8.9
"""
import numpy as np
import matplotlib.pyplot as plt
import discrete_dmp_1D
from scipy.interpolate import interp1d
import pybullet as p
import pybullet_data as pd
from openpyxl import load_workbook


# ------ Function: draw target circle
def draw_area(surface, c_x, c_y, w, color):
    from_s = [[c_x+w/2, c_y+w/2, 0], [c_x+w/2, c_y+w/2, 0], [c_x-w/2, c_y+w/2, 0], [c_x+w/2, c_y-w/2, 0]]
    to_s = [[c_x+w/2, c_y-w/2, 0], [c_x-w/2, c_y+w/2, 0], [c_x-w/2, c_y-w/2, 0], [c_x-w/2, c_y-w/2, 0]]
    for f, t in zip(from_s, to_s):
        surface.addUserDebugLine(lineFromXYZ=f, lineToXYZ=t, lineColorRGB=color, lineWidth=2)


# ------ User-adjustable parameters (suggested)
k_interpolation = 2  # multiple of interpolation, integer greater than or equal to 1
n_rbf = 500  # numbers of DMP basis functions, integer
tau = 1.0  # scaling in time
bias_1e = [0, -0.2, 0, 0, 0, 0, 0, 0]
bias_2e = [0, -1, 0.03, 0, 0, 0, 0, 0]


# ------ Load data from demonstration .xlsx
data_path = 'demonstrations/demo2023-08-21_13-21-12.xlsx'
wb = load_workbook(data_path)
mySheet = wb.active
data_length = mySheet.max_row
demo_x = np.zeros(data_length-1)
demo_y = np.zeros(data_length-1)
demo_z = np.zeros(data_length-1)
demo_q1 = np.zeros(data_length-1)
demo_q2 = np.zeros(data_length-1)
demo_q3 = np.zeros(data_length-1)
demo_q4 = np.zeros(data_length-1)
demo_finger = np.zeros(data_length-1)
for i in range(2, data_length+1):
    # demo_x
    d2 = (mySheet.cell(row=i, column=2)).value
    demo_x[i - 2] = d2
    # demo_y
    d3 = (mySheet.cell(row=i, column=3)).value
    demo_y[i - 2] = d3
    # demo_z
    d4 = (mySheet.cell(row=i, column=4)).value
    demo_z[i - 2] = d4
    # demo_q1
    d5 = (mySheet.cell(row=i, column=5)).value
    demo_q1[i - 2] = d5
    # demo_q2
    d6 = (mySheet.cell(row=i, column=6)).value
    demo_q2[i - 2] = d6
    # demo_q3
    d7 = (mySheet.cell(row=i, column=7)).value
    demo_q3[i - 2] = d7
    # demo_q4
    d8 = (mySheet.cell(row=i, column=8)).value
    demo_q4[i - 2] = d8
    # demo_finger
    d9 = (mySheet.cell(row=i, column=9)).value
    demo_finger[i - 2] = d9
# to avoid the "start equal to target" problem of classical DMP
demo_finger[0] = demo_finger[0] + 0.000001
# task segmentation
cnt = 0
for i in range(2, len(demo_finger)-1):
    if demo_finger[i] > 0 and demo_finger[i+1] == 0:
        cnt = i+5
demo_finger1 = demo_finger[:cnt]
demo_finger2 = demo_finger[cnt:]

demo_x1 = demo_x[:cnt]
demo_x2 = demo_x[cnt:]
demo_y1 = demo_y[:cnt]
demo_y2 = demo_y[cnt:]
demo_z1 = demo_z[:cnt]
demo_z2 = demo_z[cnt:]

demo_q11 = demo_q1[:cnt]
demo_q12 = demo_q1[cnt:]
demo_q21 = demo_q2[:cnt]
demo_q22 = demo_q2[cnt:]
demo_q31 = demo_q3[:cnt]
demo_q32 = demo_q3[cnt:]
demo_q41 = demo_q4[:cnt]
demo_q42 = demo_q4[cnt:]
print("------Loading demo successfully------")


# ------ Train DMPs: part1
# train demo
dmp_x1 = discrete_dmp_1D.DiscreteDMP(data_set=demo_x1, n_rbf=n_rbf)
dmp_y1 = discrete_dmp_1D.DiscreteDMP(data_set=demo_y1, n_rbf=n_rbf)
dmp_z1 = discrete_dmp_1D.DiscreteDMP(data_set=demo_z1, n_rbf=n_rbf)
dmp_q11 = discrete_dmp_1D.DiscreteDMP(data_set=demo_q11, n_rbf=n_rbf)
dmp_q21 = discrete_dmp_1D.DiscreteDMP(data_set=demo_q21, n_rbf=n_rbf)
dmp_q31 = discrete_dmp_1D.DiscreteDMP(data_set=demo_q31, n_rbf=n_rbf)
dmp_q41 = discrete_dmp_1D.DiscreteDMP(data_set=demo_q41, n_rbf=n_rbf)
dmp_finger1 = discrete_dmp_1D.DiscreteDMP(data_set=demo_finger1, n_rbf=n_rbf)
# reproduction based on DMP
x1_re, dx_re, ddx_re = dmp_x1.reproduction(start=demo_x1[0], target=demo_x1[-1]+bias_1e[0], tau=tau)
y1_re, dy_re, ddy_re = dmp_y1.reproduction(start=demo_y1[0], target=demo_y1[-1]+bias_1e[1], tau=tau)
z1_re, dz_re, ddz_re = dmp_z1.reproduction(start=demo_z1[0], target=demo_z1[-1]+bias_1e[2], tau=tau)
q11_re, dq1_re, ddq1_re = dmp_q11.reproduction(start=demo_q11[0], target=demo_q11[-1]+bias_1e[3], tau=tau)
q21_re, dq2_re, ddq2_re = dmp_q21.reproduction(start=demo_q21[0], target=demo_q21[-1]+bias_1e[4], tau=tau)
q31_re, dq3_re, ddq3_re = dmp_q31.reproduction(start=demo_q31[0], target=demo_q31[-1]+bias_1e[5], tau=tau)
q41_re, dq4_re, ddq4_re = dmp_q41.reproduction(start=demo_q41[0], target=demo_q41[-1]+bias_1e[6], tau=tau)
finger1_re, df_re, ddf_re = dmp_finger1.reproduction(start=demo_finger1[0], target=demo_finger1[-1]+bias_1e[7], tau=tau)
print("------Training DMPs successfully------")
# ------ Interpolating the reproduced trajectory makes the motion control smoother
# constructing time series
n_data_o = np.linspace(0, len(x1_re), len(x1_re))
target_n = len(x1_re) * k_interpolation
n_new = np.linspace(0, len(x1_re), target_n)
# interpolate x y z
interp_func_x1 = interp1d(n_data_o, x1_re, kind='cubic')
target_x1 = interp_func_x1(n_new)
interp_func_y1 = interp1d(n_data_o, y1_re, kind='cubic')
target_y1 = interp_func_y1(n_new)
interp_func_z1 = interp1d(n_data_o, z1_re, kind='cubic')
target_z1 = interp_func_z1(n_new)
# interpolate q1 q2 q3 q4
interp_func_q11 = interp1d(n_data_o, q11_re, kind='cubic')
target_q11 = interp_func_q11(n_new)
interp_func_q21 = interp1d(n_data_o, q21_re, kind='cubic')
target_q21 = interp_func_q21(n_new)
interp_func_q31 = interp1d(n_data_o, q31_re, kind='cubic')
target_q31 = interp_func_q31(n_new)
interp_func_q41 = interp1d(n_data_o, q41_re, kind='cubic')
target_q41 = interp_func_q41(n_new)
# interpolate finger
interp_func_f1 = interp1d(n_data_o, finger1_re, kind='cubic')
target_finger1 = interp_func_f1(n_new)
print("------Interpolating reproduction successfully: part 1------")


# ------ Train DMPs: part2
# train demo
dmp_x2 = discrete_dmp_1D.DiscreteDMP(data_set=demo_x2, n_rbf=n_rbf)
dmp_y2 = discrete_dmp_1D.DiscreteDMP(data_set=demo_y2, n_rbf=n_rbf)
dmp_z2 = discrete_dmp_1D.DiscreteDMP(data_set=demo_z2, n_rbf=n_rbf)
dmp_q12 = discrete_dmp_1D.DiscreteDMP(data_set=demo_q12, n_rbf=n_rbf)
dmp_q22 = discrete_dmp_1D.DiscreteDMP(data_set=demo_q22, n_rbf=n_rbf)
dmp_q32 = discrete_dmp_1D.DiscreteDMP(data_set=demo_q32, n_rbf=n_rbf)
dmp_q42 = discrete_dmp_1D.DiscreteDMP(data_set=demo_q42, n_rbf=n_rbf)
dmp_finger2 = discrete_dmp_1D.DiscreteDMP(data_set=demo_finger2, n_rbf=n_rbf)
# reproduction based on DMP
x2_re, dx2_re, ddx2_re = dmp_x2.reproduction(start=demo_x2[0]+bias_1e[0], target=demo_x2[-1]+bias_2e[0], tau=tau)
y2_re, dy2_re, ddy2_re = dmp_y2.reproduction(start=demo_y2[0]+bias_1e[1], target=demo_y2[-1]+bias_2e[1], tau=tau)
z2_re, dz2_re, ddz2_re = dmp_z2.reproduction(start=demo_z2[0]+bias_1e[2], target=demo_z2[-1]+bias_2e[2], tau=tau)
q12_re, dq12_re, ddq12_re = dmp_q12.reproduction(start=demo_q12[0]+bias_1e[3], target=demo_q12[-1]+bias_2e[3], tau=tau)
q22_re, dq22_re, ddq22_re = dmp_q22.reproduction(start=demo_q22[0]+bias_1e[4], target=demo_q22[-1]+bias_2e[4], tau=tau)
q32_re, dq32_re, ddq32_re = dmp_q32.reproduction(start=demo_q32[0]+bias_1e[5], target=demo_q32[-1]+bias_2e[5], tau=tau)
q42_re, dq42_re, ddq42_re = dmp_q42.reproduction(start=demo_q42[0]+bias_1e[6], target=demo_q42[-1]+bias_2e[6], tau=tau)
finger2_re, df2_re, ddf2_re = dmp_finger2.reproduction(start=demo_finger2[0]+bias_1e[7], target=demo_finger2[-1]+bias_2e[7], tau=tau)
print("------Training DMPs successfully------")
# ------ Interpolating the reproduced trajectory makes the motion control smoother
# constructing time series
n_data_o = np.linspace(0, len(x2_re), len(x2_re))
target_n = len(x2_re) * k_interpolation
n_new = np.linspace(0, len(x2_re), target_n)
# interpolate x y z
interp_func_x2 = interp1d(n_data_o, x2_re, kind='cubic')
target_x2 = interp_func_x2(n_new)
interp_func_y2 = interp1d(n_data_o, y2_re, kind='cubic')
target_y2 = interp_func_y2(n_new)
interp_func_z2 = interp1d(n_data_o, z2_re, kind='cubic')
target_z2 = interp_func_z2(n_new)
# interpolate q1 q2 q3 q4
interp_func_q12 = interp1d(n_data_o, q12_re, kind='cubic')
target_q12 = interp_func_q12(n_new)
interp_func_q22 = interp1d(n_data_o, q22_re, kind='cubic')
target_q22 = interp_func_q22(n_new)
interp_func_q32 = interp1d(n_data_o, q32_re, kind='cubic')
target_q32 = interp_func_q32(n_new)
interp_func_q42 = interp1d(n_data_o, q42_re, kind='cubic')
target_q42 = interp_func_q42(n_new)
# interpolate finger
interp_func_f2 = interp1d(n_data_o, finger2_re, kind='cubic')
target_finger2 = interp_func_f2(n_new)
print("------Interpolating reproduction successfully: part 2------")


print("------Start simulation------")
# ------ Connect to the pybullet server and load the environment
p.connect(p.GUI)
p.setAdditionalSearchPath(pd.getDataPath())
# load models
pandaUid = p.loadURDF("franka_panda/panda.urdf", useFixedBase=True)
tableUid = p.loadURDF("table/table.urdf", basePosition=[0.5, 0, -0.65])
table2Uid = p.loadURDF("table/table.urdf", basePosition=[0.5, -1, -0.65])
table3Uid = p.loadURDF("table/table.urdf", basePosition=[0.5, 1, -0.65])
# trayUid = p.loadURDF("tray/traybox.urdf", basePosition=[1, 0, 0], baseOrientation=[0, 1, 0, 0], globalScaling=0.5)
objectUid = p.loadURDF("random_urdfs/000/000.urdf", basePosition=[0.7, 0, 0.2])
object2Uid = p.loadURDF("random_urdfs/000/000.urdf", basePosition=[0.4, 0, 0.2])
object3Uid = p.loadURDF("random_urdfs/000/000.urdf", basePosition=[0.7, -0.2, 0.2])
# set gravity and camera
p.setGravity(0, 0, -9.81)
p.resetDebugVisualizerCamera(cameraDistance=1, cameraYaw=0,
                             cameraPitch=-45, cameraTargetPosition=[0.55, -0.35, 0.2])
# draw the target area
draw_area(p, 0.5, 0.45, 0.2, [255, 255, 255])
draw_area(p, 0.5, -0.5, 0.2, [255, 0, 0])


# ------ Simulation start
skill = "pick"
k = 0
joint_index = 7
while True:
    p.configureDebugVisualizer(p.COV_ENABLE_SINGLE_STEP_RENDERING)
    p.stepSimulation()
    # press q to break simulation
    keys = p.getKeyboardEvents()
    if ord("q") in keys:
        break

    if skill == "pick":
        # reproduction
        target_xyz = [target_x1[k], target_y1[k], target_z1[k]]
        target_q = [target_q11[k], target_q21[k], target_q31[k], target_q41[k]]
        jointAngles = p.calculateInverseKinematics(pandaUid, joint_index, target_xyz, target_q, maxNumIterations=1000)
        # control
        for i in range(7):
            p.setJointMotorControl2(pandaUid, i, p.POSITION_CONTROL, jointAngles[i])
        p.setJointMotorControl2(pandaUid, 9, p.POSITION_CONTROL, target_finger1[k])
        p.setJointMotorControl2(pandaUid, 10, p.POSITION_CONTROL, target_finger1[k])
        # status
        k = k + 1
        if k == len(target_x1):
            skill = "place"
            k = 0

    if skill == "place":
        # reproduction
        target_xyz = [target_x2[k], target_y2[k], target_z2[k]]
        target_q = [target_q12[k], target_q22[k], target_q32[k], target_q42[k]]
        jointAngles = p.calculateInverseKinematics(pandaUid, joint_index, target_xyz, target_q, maxNumIterations=1000)
        # control
        for i in range(7):
            p.setJointMotorControl2(pandaUid, i, p.POSITION_CONTROL, jointAngles[i])
        p.setJointMotorControl2(pandaUid, 9, p.POSITION_CONTROL, target_finger2[k])
        p.setJointMotorControl2(pandaUid, 10, p.POSITION_CONTROL, target_finger2[k])
        # status
        k = k + 1
        if k == len(target_x2):
            break
print("------Simulation over------")
